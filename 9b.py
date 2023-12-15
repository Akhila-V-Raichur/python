from openpyxl import Workbook
from openpyxl.styles import Font

wb=Workbook()
sheet=wb.active
sheet.title="Language"
wb.create_sheet(title="Capital")

lang=["Kannada","Telugu","Tamil"]
state=["Karnataka","Telangana","Tamil nadu"]
capital=["Bengaluru","Hyderabad","Chennai"]
code=['KA','TN','TS']

sheet.cell(row=1,column=1).value="State"
sheet.cell(row=1,column=2).value="Language"
sheet.cell(row=1,column=3).value="Code"

ft=Font(bold=True)
for row in sheet["A1:C1"]:
    for cell in row:
        cell.font=ft

for i in range(2,5):
    sheet.cell(row=i,column=1).value=state[i-2]
    sheet.cell(row=i,column=2).value=lang[i-2]
    sheet.cell(row=i,column=3).value=code[i-2]

wb.save('demo.xlsx')

sheet=wb["Capital"]

sheet.cell(row=1,column=1).value="State"
sheet.cell(row=1,column=2).value="Capital"
sheet.cell(row=1,column=3).value="Code"

ft=Font(bold=True)
for row in sheet["A1:C1"]:
    for cell in row:
        cell.font=ft

for i in range(2,5):
    sheet.cell(row=i,column=1).value=state[i-2]
    sheet.cell(row=i,column=2).value=capital[i-2]
    sheet.cell(row=i,column=3).value=code[i-2]

wb.save('demo.xlsx')
 
srchcode=input("enter the state code to find the capital:")
for i in range(2,5):
    data=sheet.cell(row=i,column=3).value
    if data==srchcode:
        print("Corresponding capital for code", srchcode, "is", sheet.cell(row = i, column = 2).value)

sheet=wb["Language"]

srchode=input("enter the state code to find the language:")
for i in range(2,5):
    data=sheet.cell(row=i,column=3).value
    if data==srchcode:
        print("Corresponding capital for code", srchcode, "is", sheet.cell(row = i, column = 2).value)

wb.close()